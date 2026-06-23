from rest_framework import generics, permissions, status
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from django.db.models import Q
from .authentication import BearerTokenAuthentication
from .models import Lead
from .serializers import LeadCreateSerializer, LeadAdminSerializer
import csv
from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r')

def sanitize_cell(value):
    if value is None:
        return ''

    if isinstance(value, str) and value.lstrip().startswith(FORMULA_PREFIXES):
        return "'" + value

    return value

def _apply_filters(queryset, request):
    search   = request.query_params.get('search',   '').strip()
    segment  = request.query_params.get('segment',  '').strip()
    interest = request.query_params.get('interest', '').strip()
    city     = request.query_params.get('city',     '').strip()   # NOVO

    if search:
        queryset = queryset.filter(
            Q(nome_completo__icontains=search) | Q(email__icontains=search)
        )
    if segment:
        queryset = queryset.filter(segmento=segment)
    if interest:
        from .serializers import INTEREST_API_TO_MODEL
        model_value = INTEREST_API_TO_MODEL.get(interest)
        queryset = queryset.filter(como_ajudar=model_value) if model_value else queryset.none()
    if city:                                                        # NOVO
        queryset = queryset.filter(cidade_estado__icontains=city)  # NOVO

    return queryset
class LeadExportXlsxView(APIView):
    authentication_classes = [BearerTokenAuthentication, SessionAuthentication, BasicAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not XLSX_AVAILABLE:
            return Response(
                {'erro': 'Exportação XLSX não disponível. Instale openpyxl.'},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

        qs = _apply_filters(Lead.objects.all(), request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads AccessSim"

        headers = [
            'Nome completo', 'E-mail', 'Organização', 'Cargo/Função',
            'Telefone', 'Cidade/Estado', 'Segmento', 'Como posso ajudar',
            'Mensagem', 'Consentimento', 'Origem', 'Criado em',
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A2F4A")

        for col, header in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, lead in enumerate(qs, 2):
            row = [
                sanitize_cell(lead.nome_completo),
                sanitize_cell(lead.email),
                sanitize_cell(lead.organizacao),
                sanitize_cell(lead.cargo_funcao),
                sanitize_cell(lead.telefone),
                sanitize_cell(lead.cidade_estado),
                lead.get_segmento_display(),
                lead.get_como_ajudar_display(),
                sanitize_cell(lead.mensagem or ''),
                'Sim' if lead.consentimento else 'Não',
                sanitize_cell(lead.origem),
                lead.criado_em.strftime('%d/%m/%Y %H:%M'),
            ]
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        col_widths = [30, 35, 30, 25, 20, 20, 20, 25, 50, 15, 15, 20]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="leads_accesssim.xlsx"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        wb.save(response)
        return response

class LeadCreateView(generics.CreateAPIView):
    queryset = Lead.objects.all()
    serializer_class = LeadCreateSerializer
    permission_classes = [permissions.AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            # LeadForm.jsx espera erros em err.erros, não no formato padrão do DRF
            return Response({'erros': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        # LeadForm.jsx exibe res.mensagem na tela de sucesso
        return Response(
            {'mensagem': 'Recebemos sua mensagem! Nossa equipe entrará em contato em breve.'},
            status=status.HTTP_201_CREATED,
        )


class LeadListView(generics.ListAPIView):
    queryset = Lead.objects.all().order_by('-criado_em')
    serializer_class = LeadAdminSerializer
    authentication_classes = [
        BearerTokenAuthentication,
        SessionAuthentication,
        BasicAuthentication,
    ]
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        queryset = _apply_filters(self.get_queryset(), request)
        serializer = self.get_serializer(queryset, many=True)

        return Response({
            'leads': serializer.data,
            'total': queryset.count(),
        })


class LeadDetailView(generics.RetrieveAPIView):
    queryset = Lead.objects.all()
    serializer_class = LeadAdminSerializer
    authentication_classes = [BearerTokenAuthentication, SessionAuthentication, BasicAuthentication]
    permission_classes = [permissions.IsAuthenticated]

class AdminLoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        username = request.data.get('username', '').strip()
        password = request.data.get('password', '') 
        user = authenticate(request, username=username, password=password)
        if user is None:
            return Response({'erro': 'Usuário ou senha inválidos.'}, status=status.HTTP_401_UNAUTHORIZED)
        token, _ = Token.objects.get_or_create(user=user)
        return Response({'token': token.key, 'username': user.username})

class LeadExportCsvView(APIView):
    authentication_classes = [BearerTokenAuthentication, SessionAuthentication, BasicAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        queryset = _apply_filters(
            Lead.objects.all().order_by('-criado_em'), request
        )

        response = HttpResponse(
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = (
            'attachment; filename="leads.csv"'
        )
        response['Access-Control-Expose-Headers'] = (
            'Content-Disposition'
        )

        # Faz o Excel reconhecer corretamente os caracteres UTF-8.
        response.write('\ufeff')

        writer = csv.writer(response)

        writer.writerow([
            'Nome completo',
            'E-mail',
            'Organização',
            'Cargo/Função',
            'Telefone',
            'Cidade/Estado',
            'Segmento',
            'Como posso ajudar',
            'Mensagem',
            'Consentimento',
            'Origem',
            'Criado em',
        ])

        for lead in queryset:
            writer.writerow([
                sanitize_cell(lead.nome_completo),
                sanitize_cell(lead.email),
                sanitize_cell(lead.organizacao),
                sanitize_cell(lead.cargo_funcao),
                sanitize_cell(lead.telefone),
                sanitize_cell(lead.cidade_estado),
                sanitize_cell(lead.get_segmento_display()),
                sanitize_cell(lead.get_como_ajudar_display()),
                sanitize_cell(lead.mensagem),
                'Sim' if lead.consentimento else 'Não',
                sanitize_cell(lead.origem),
                lead.criado_em.strftime('%d/%m/%Y %H:%M'),
            ])

        return response